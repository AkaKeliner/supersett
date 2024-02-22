# Licensed to the Apache Software Foundation (ASF) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The ASF licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
import datetime
import decimal
import io
from typing import Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Side, Border
from superset.config import DATE_FORMATS
import locale


def df_to_excel(df: pd.DataFrame, additional_data=None, **kwargs: Any) -> Any:

    # timezones are not supported
    for column in df.select_dtypes(include=["datetimetz"]).columns:
        df[column] = df[column].astype(str)

    # pylint: disable=abstract-class-instantiated
    # with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
    #     df.to_excel(writer, **kwargs)
    try:
        output = add_width_normalization(df, additional_data, **kwargs)
    except Exception as e:
        print(f"catch error in df_to_excel:\n {e.args}")
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, **kwargs)
    return output.getvalue()


def add_width_normalization(
    df: pd.DataFrame, additional_data=None, **kwargs: Any
) -> Any:
    output = io.BytesIO()

    # Записываем DataFrame в файл Excel с помощью openpyxl
    # with pd.ExcelWriter(output, engine="openpyxl") as writer:
    #     df.to_excel(writer, index=False, **kwargs)
    #
    #     # Получаем объект workbook и worksheet
    #     workbook = writer.book
    #     worksheet = writer.sheets["Sheet1"]
    #
    #     # Устанавливаем максимальную ширину столбцов
    #     for col in worksheet.columns:
    #         max_length = 0
    #         column = col[0].column_letter
    #         for cell in col:
    #             try:
    #                 if len(str(cell.value)) > max_length:
    #                     max_length = len(cell.value)
    #             except:
    #                 pass
    #         adjusted_width = (
    #             max_length + 2
    #         ) * 1.2  # 1.2 - коэффициент масштабирования для учета ширины символов
    #         worksheet.column_dimensions[column].width = min(350, adjusted_width)
    # Создаем объект Workbook
    workbook = Workbook()

    # Создаем объект Worksheet
    worksheet = workbook.active

    # Записываем данные DataFrame в Excel файл с использованием openpyxl
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)
        # Словарь для отслеживания начальных индексов строк групп одинаковых значений
    merge_groups = {}

    # Проверяем каждый столбец на наличие групп одинаковых значений и объединяем их
    for col_idx, column in enumerate(df.columns, start=1):
        prev_value = None
        start_row = None
        for row_idx, cell in enumerate(
            worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx), start=2
        ):
            value = cell[0].value
            if value == prev_value:
                if start_row is None:
                    start_row = row_idx - 1
                end_row = row_idx
            elif start_row is not None:
                merge_groups[(col_idx, start_row)] = end_row
                start_row = None
            prev_value = value
        if start_row is not None:
            merge_groups[(col_idx, start_row)] = end_row

    # Объединяем ячейки с одинаковыми значениями
    for (col_idx, start_row), end_row in merge_groups.items():
        worksheet.merge_cells(
            start_row=start_row,
            end_row=end_row,
            start_column=col_idx,
            end_column=col_idx,
        )

    # Устанавливаем максимальную ширину столбцов
    for col in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = min(350, (max_length + 2) * 1.2)
        worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

    # Сохраняем Workbook в файл

    # Устанавливаем выравнивание для переноса текста
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True)

    # Автоматически подгоняем высоту строк
    for row in worksheet.iter_rows():
        max_height = 0
        for cell in row:
            if isinstance(cell.value, str):
                lines = cell.value.split("\n")
                height = max(len(line) // 100 + 1 for line in lines)
                if height > max_height:
                    max_height = height
            # if cell.value:
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            cell.border = border
        worksheet.row_dimensions[row[0].row].height = (
            max_height * 15
        )  # 15 - это высота одной строки (подгоните под ваш шрифт)

    # Сохраняем Workbook в файл
    columns_format = get_formats_column(
        worksheet, column_configs=additional_data.get("column_config")
    )
    columns_funcs = convert_formats_to_funcs(columns_format)
    merged_cells_range = worksheet.merged_cells.ranges
    for row in worksheet.iter_rows(
        min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row
    ):
        for cell in row:
            if cell.value is None:
                continue
            header_column = worksheet.cell(row=1, column=cell.column)
            if merged_cells_range:
                merged_cells_range_tuple = [
                    merged.bounds[0::2] for merged in merged_cells_range
                ]
                for i, merged in enumerate(merged_cells_range_tuple):
                    if merged[0] <= cell.column <= merged[1]:
                        header_column = worksheet.cell(row=1, column=merged[0])
                        break
            cell_value = cell.value
            if isinstance(cell_value, (int, float, datetime.datetime, datetime.date, decimal.Decimal)):
                cell_value=columns_format[header_column.value](float(cell.value))
                cell.value = cell_value
    workbook.save(output)

    return output


def convert_formats_cell(worksheet: Worksheet, columns_func):
    merged_cells_range = worksheet.merged_cells.ranges
    for row in worksheet.iter_rows(
        min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row
    ):
        for cell in row:
            if cell.value is None:
                continue
            header_column = worksheet.cell(row=1, column=cell.column)
            if merged_cells_range:
                merged_cells_range_tuple = [
                    merged.bounds[0::2] for merged in merged_cells_range
                ]
                for i, merged in enumerate(merged_cells_range_tuple):
                    if merged[0] <= cell.column <= merged[1]:
                        header_column = worksheet.cell(row=1, column=merged[0])
                        break
            cell_value = cell.value
            if isinstance(cell_value, (int, float, datetime.datetime, datetime.date)):
                cell.value = columns_func[header_column.value](cell.value)
    return worksheet


def get_formats_column(worksheet, column_configs: dict = None):
    columns_format = {}
    for column in worksheet[1]:
        if column_configs is not None and column.value not in column_configs.keys():
            columns_format[column.value] = ""
        else:
            columns_format[column.value] = column_configs[column.value].get(
                "d3NumberFormat"
            )
    return columns_format


def convert_formats_to_funcs(column_configs: dict):
    for k, v in column_configs.items():
        if v in DATE_FORMATS:
            column_configs[k] = make_format_date_ext(v)
        elif v == "":
            column_configs[k] = lambda value: value
        else:
            column_configs[k] = make_format_number(v)
    return column_configs


def make_format_date_ext(time_grain):
    return lambda dttm: datetime.strptime(
        dttm.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S"
    ).strftime(time_grain)


def make_format_number(specifier):
    locale.setlocale(locale.LC_ALL, "C")
    if "%" in specifier:
        decimals = int(specifier.split(".")[1][0])
        return lambda value: format_number_with_percentage(
            value, decimals
        )  # Установка локали для корректного форматирования чисел
    elif "s" in specifier:
        decimals = int(
            specifier.split(".")[1][0]
        )  # Извлекаем количество знаков после запятой из формата
        return lambda value: format_number_with_suffix(value, decimals, specifier)
    elif "r" in specifier:
        decimals = int(specifier.split(".")[1][0])
        return lambda value: format_number_with_rounding(value, decimals)
    elif "+" in specifier:
        return lambda value: "{:+,}".format(value)
    elif "$" in specifier:
        return lambda value: "₽{:,.2f}".format(value)
    return lambda value: format(value, specifier)


def format_number_with_percentage(number, decimals):
    rounded_number = round(
        number * 100, decimals
    )  # Преобразуем число в проценты и округляем
    if decimals == 1:
        # Особый случай для формата .1%
        if int(rounded_number) == rounded_number:
            return f"{int(rounded_number)}%"  # Если число целое, выводим его без десятичной части
    return f"{rounded_number:.{decimals}f}%"


def format_number_with_suffix(number, decimals, specifier=None):
    suffixes = ["", "тыс.", "млн.", "млрд.", "трлн."]  # Суффиксы для кратных единиц
    index = 0
    while abs(number) >= 1000 and index < len(suffixes) - 1:
        index += 1
        number /= 1000.0
    rounded_number = round(number, decimals)
    suf = ""
    if specifier and "$" in specifier:
        suf = "₽"
    # Округляем число с учетом указанного количества знаков после запятой
    if int(rounded_number) == 0:
        return f"{suf}0"
    elif int(rounded_number) == rounded_number:
        return f"{suf}{int(rounded_number)} {suffixes[index]}"  # Если число округляется до целого, преобразуем его в целое число
    else:
        return f"{suf}{rounded_number:.{decimals}f} {suffixes[index]}"  # Иначе, округляем число и добавляем суффикс


def format_number_with_rounding(number, decimals):
    rounded_number = round(number, decimals)  # Округляем число
    return f"{rounded_number:.{decimals}f}"  # Форматируем число с заданным количеством десятичных знаков
